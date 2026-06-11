"""Aggregate Distribution Rewards per prime agent from the DR workbooks.

Inputs (fetched from their canonical public URLs, cached in /tmp;
``--local`` falls back to the snapshot copies in ``preparation/``):

* DR results — ``dr_comparison_2026.xlsx`` on GitHub (settle-dr-dune repo;
  identical to the local ``dr_results.xlsx`` snapshot). The ``Soter Data``
  sheet is OpenMSC's DR-repo output; ``Spark`` / ``Amatsu`` are
  counterparty calculations kept for diffing. Ref codes appear either
  bare (``128``) or decorated with a farm/token suffix (``0 (sUSDS)``,
  ``1 (stUSDS)``) — the integer prefix is the referral code.
* DR payouts — the published Google sheet (CSV endpoint; identical to the
  local ``dr_payouts.xlsx`` snapshot): referral code → Partner/Prime
  mapping plus the historical monthly payouts. The prime is the
  parenthesised suffix when it names a known prime
  (``Spark.lend (Spark)`` → Spark), else the label/prefix itself
  (``Spark``, ``Skybase``, ``Grove (Maple)`` → Grove, ``untagged``).

Usage:
    python3 preparation/scripts/dr_aggregate.py --prime spark
    python3 preparation/scripts/dr_aggregate.py --prime skybase --sheet "Soter Data"
    python3 preparation/scripts/dr_aggregate.py --list-primes

Prints per-month totals (2026-01..05), Jan–Apr / Jan–May sums, the
matching past payouts from ``dr_payouts.xlsx`` (history ends 2026-03),
the contributing referral codes, and a ready-to-paste markdown table.
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import sys
import time
import urllib.request
from decimal import Decimal
from pathlib import Path

import openpyxl

_PREP = Path(__file__).resolve().parent.parent
# Local snapshots (offline fallback via --local). ⚠️ These are point-in-time
# copies (last refreshed 2026-06-11) of the remote sources below — the
# remote CSV/xlsx update upstream while the snapshots stay frozen, so
# ``--local`` and the default remote mode can legitimately diverge once the
# upstream files change (e.g. when Apr/May 2026 payout columns are added).
RESULTS_XLSX = _PREP / "dr_results.xlsx"
PAYOUTS_XLSX = _PREP / "dr_payouts.xlsx"
# Canonical public sources.
RESULTS_URL = ("https://raw.githubusercontent.com/stablewatch-io/settle-dr-dune/"
               "main/dune-results/dr_comparison_2026.xlsx")
PAYOUTS_CSV_URL = ("https://docs.google.com/spreadsheets/d/e/"
                   "2PACX-1vR-dLvndU-DM1j_8gxYIhfYOtoIgyEJ9Jg5R0RcV-ZRVGdOJdmwIysO4P9yfacw-CkBGJjXPgwbC6WB/"
                   "pub?gid=753291318&single=true&output=csv")

USE_LOCAL = False  # set by --local

MONTHS = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05"]

# Published-CSV month headers ("Jan 2026") → our keys ("2026-01").
_MONTH_NAMES = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}


def _month_key(header: str) -> str | None:
    """``'Jan 2026'`` or ``'2026-01-01 00:00:00'`` → ``'2026-01'``; else None."""
    h = str(header).strip()
    if re.fullmatch(r"\d{4}-\d{2}.*", h):
        return h[:7]
    m = re.fullmatch(r"([A-Za-z]{3})\w* (\d{4})", h)
    if m and m.group(1) in _MONTH_NAMES:
        return f"{m.group(2)}-{_MONTH_NAMES[m.group(1)]:02d}"
    return None


_CACHE_TTL_HOURS = 24


def _fetch(url: str, cache_name: str) -> bytes:
    """Download ``url``, caching in /tmp. A cache older than
    ``_CACHE_TTL_HOURS`` is re-fetched; a fresh-enough cache prints a
    notice so a stale read is never silent."""
    cache = Path("/tmp") / cache_name
    if cache.exists():
        age_h = (time.time() - cache.stat().st_mtime) / 3600
        if age_h <= _CACHE_TTL_HOURS:
            print(f"[cache] using /tmp/{cache_name} ({age_h:.1f}h old; "
                  f"delete it to force re-fetch)", file=sys.stderr)
            return cache.read_bytes()
        print(f"[cache] /tmp/{cache_name} is {age_h:.0f}h old — re-fetching",
              file=sys.stderr)
    with urllib.request.urlopen(url) as r:
        data = r.read()
    cache.write_bytes(data)
    return data


def _payout_rows() -> list[list[str]]:
    """[[Referral Code, Partner/Prime, <month cells…>], …] incl. header."""
    if USE_LOCAL:
        ws = openpyxl.load_workbook(PAYOUTS_XLSX, read_only=True, data_only=True).active
        return [[("" if c is None else str(c)) for c in r]
                for r in ws.iter_rows(values_only=True)]
    raw = _fetch(PAYOUTS_CSV_URL, "dr_payouts_pub.csv").decode("utf-8")
    return [list(r) for r in csv.reader(io.StringIO(raw))]

# ``Spark.lend (Spark)`` → Spark; bare labels map to themselves.
_PAREN = re.compile(r"\(([^)]+)\)\s*$")

# The parenthesised suffix is the prime ONLY when it names a known prime
# (``summerfi (Skybase)`` → Skybase). ``Grove (Maple)`` breaks the
# ``partner (Prime)`` convention — Maple is Grove's partner venue, so the
# prime is the PREFIX there (confirmed 2026-06-11; Grove's March DR then
# matches the MSC#7 forum demand sub-total).
_KNOWN_PRIMES = {"spark", "grove", "skybase", "keel", "obex"}


def _prime_of(label: str) -> str:
    label = label.strip()
    m = _PAREN.search(label)
    if m and m.group(1).strip().lower() in _KNOWN_PRIMES:
        return m.group(1).strip()
    if m:  # ``Grove (Maple)`` — strip the partner suffix, keep the prefix
        return label[: m.start()].strip()
    return label


def _code_of(ref_code: str) -> str:
    """``'0 (sUSDS)'`` → ``'0'``; ``'128'`` → ``'128'``; passthrough else."""
    return str(ref_code).strip().split(" ")[0].split(".")[0]


def _dec(v) -> Decimal:
    if v is None or str(v).strip() == "":
        return Decimal("0")
    return Decimal(str(v).replace(",", ""))  # published CSV uses 1,234.50


def load_code_to_prime() -> dict[str, str]:
    rows = _payout_rows()
    out: dict[str, str] = {}
    for r in rows[1:]:
        if len(r) < 2 or not str(r[0]).strip() or not str(r[1]).strip():
            continue
        out[_code_of(r[0])] = _prime_of(str(r[1]))
    return out


def load_results(sheet: str) -> list[tuple[str, dict[str, Decimal]]]:
    """[(ref_code_label, {month: amount})] from the DR results workbook."""
    if USE_LOCAL:
        wb = openpyxl.load_workbook(RESULTS_XLSX, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(
            io.BytesIO(_fetch(RESULTS_URL, "dr_comparison_2026.xlsx")),
            read_only=True, data_only=True,
        )
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {m: hdr.index(m) for m in MONTHS if m in hdr}
    out = []
    for r in rows[1:]:
        if r[0] is None or str(r[0]).strip() == "":
            continue
        out.append((str(r[0]).strip(), {m: _dec(r[i]) for m, i in col.items()}))
    return out


def load_past_payouts(prime: str, code_to_prime: dict[str, str]) -> dict[str, Decimal]:
    """{month: payout} summed over the prime's codes, for the 2026 month
    columns present in the payouts sheet (history currently ends 2026-03)."""
    rows = _payout_rows()
    hdr = rows[0]
    col_by_month: dict[str, int] = {}
    for i, c in enumerate(hdr):
        key = _month_key(c)
        if key in MONTHS:
            col_by_month[key] = i
    out = {m: Decimal("0") for m in col_by_month}
    for r in rows[1:]:
        if not str(r[0]).strip():
            continue
        if code_to_prime.get(_code_of(r[0]), "").lower() != prime.lower():
            continue
        for m, i in col_by_month.items():
            if i < len(r):
                out[m] += _dec(r[i])
    return out


def aggregate(prime: str, sheet: str) -> None:
    code_to_prime = load_code_to_prime()
    results = load_results(sheet)

    totals = {m: Decimal("0") for m in MONTHS}
    codes_used: set[str] = set()
    unmapped: set[str] = set()
    for label, by_month in results.items() if isinstance(results, dict) else results:
        code = _code_of(label)
        mapped = code_to_prime.get(code)
        if mapped is None:
            unmapped.add(code)
            continue
        if mapped.lower() != prime.lower():
            continue
        codes_used.add(code)
        for m, v in by_month.items():
            totals[m] += v

    payouts = load_past_payouts(prime, code_to_prime)

    print(f"Distribution Rewards — {prime} (sheet: {sheet!r})")
    print(f"contributing referral codes: {sorted(codes_used, key=lambda c: (len(c), c))}")
    if unmapped:
        print(f"WARNING — ref codes in results with no payout mapping (excluded): "
              f"{sorted(unmapped, key=lambda c: (len(c), c))}")
    print()
    print(f"{'Month':<10} {'OpenMSC rev':>14} {'Past payout':>14}")
    for m in MONTHS:
        po = payouts.get(m)
        po_s = f"{po:>14,.2f}" if po is not None else f"{'N/A':>14}"
        print(f"{m:<10} {totals[m]:>14,.2f} {po_s}")
    jan_apr = sum(totals[m] for m in MONTHS[:4])
    jan_may = sum(totals[m] for m in MONTHS)
    po_jan_apr = sum(payouts[m] for m in MONTHS[:4] if m in payouts)
    print(f"{'Σ Jan-Apr':<10} {jan_apr:>14,.2f} {po_jan_apr:>14,.2f}  "
          f"(payout Σ covers {sorted(payouts)} only)")
    print(f"{'Σ Jan-May':<10} {jan_may:>14,.2f}")

    print("\nmarkdown cells (OpenMSC rev | Past payout):")
    for m in MONTHS:
        po = payouts.get(m)
        po_s = f"{po:,.0f}" if po is not None else "N/A"
        print(f"  {m}: {totals[m]:,.0f} | {po_s}")
    print(f"  Total (Jan to Apr): {jan_apr:,.0f}")
    print(f"  Total (Jan to May): {jan_may:,.0f}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--prime", help="prime agent (spark, grove, skybase, keel, …)")
    ap.add_argument("--sheet", default="Soter Data",
                    help="dr_results.xlsx sheet to aggregate (default: Soter Data)")
    ap.add_argument("--list-primes", action="store_true",
                    help="list primes found in the payout mapping and exit")
    ap.add_argument("--local", action="store_true",
                    help="use the local snapshot files in preparation/ instead "
                         "of the canonical public URLs")
    args = ap.parse_args()
    global USE_LOCAL
    USE_LOCAL = args.local

    if args.list_primes:
        mapping = load_code_to_prime()
        by_prime: dict[str, list[str]] = {}
        for code, prime in mapping.items():
            by_prime.setdefault(prime, []).append(code)
        for prime in sorted(by_prime):
            codes = sorted(by_prime[prime], key=lambda c: (len(c), c))
            print(f"{prime:<10} {codes}")
        return 0

    if not args.prime:
        ap.error("--prime is required (or use --list-primes)")
    aggregate(args.prime, args.sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
