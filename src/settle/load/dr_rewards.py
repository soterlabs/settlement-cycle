"""Distribution Rewards (DR) per prime / ref code.

Sourced from the ``settle-dr-dune`` submodule's HyperSync reconciliation
workbook (``hypersync-results/dr_comparison_hypersync.xlsx``, ``Soter by Ref
Code`` tab). Each row is a ``ref_code`` with one DR-USD column per month.

Unlike the retired Dune workbook (``dune-results/dr_comparison_latest.xlsx``,
``Summary`` tab) this sheet is FLAT: it carries no ``group`` column and no
per-group ``Total`` row, so it says nothing about which prime owns a code.
Attribution therefore lives in ``config/dr_ref_codes.yaml`` in this repo, and
a group total is the sum of that prime's codes. For a requested (prime,
month) we read that month's column for the prime's codes and return the
per-ref-code breakdown plus the summed total.

A ref code present in the workbook but absent from the config is logged at
ERROR with its dollar amount and excluded — new codes cannot silently vanish
from a settlement, nor be silently attributed to the wrong prime.

Reporting-time enrichment only: ``enrich_with_dr`` populates
``MonthlyPnL.distribution_rewards`` (+ keeps the ``monthly_pnl`` invariant)
and ``MonthlyPnL.dr_breakdown`` for the summary.md "DR per ref code" table.
The provenance-patch counterpart lives in ``writer.refresh_dr_only``.

Graceful by design — returns ``None`` (and leaves the pnl unchanged) when
the submodule/workbook is absent, the prime has no DR group, or the month
isn't in the sheet. So runs without the submodule initialised still work.
"""

from __future__ import annotations

import dataclasses
import functools
import logging
from decimal import Decimal
from pathlib import Path

log = logging.getLogger(__name__)

_WORKBOOK_REL = "settle-dr-dune/hypersync-results/dr_comparison_hypersync.xlsx"
_SHEET = "Soter by Ref Code"
_REF_CODE_MAP_REL = "config/dr_ref_codes.yaml"


def _repo_root() -> Path:
    # src/settle/load/dr_rewards.py -> parents[3] is the repo root.
    return Path(__file__).resolve().parents[3]


def _dec(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    return Decimal(str(v))


def _dr_retired_from() -> dict[str, str]:
    """``{prime_id: 'YYYY-MM'}`` — first month in which the prime earns NO DR.

    From ``config/dr_ref_codes.yaml → retired_from``. Values are parsed and
    re-rendered so a malformed month ('2026-8', a full date) can't silently
    break the lexicographic comparison in ``load_dr`` — the same treatment
    ``GarConfig`` gives its bounds.

    Deliberately NOT cached, even though it re-reads the same small yaml as
    ``_ref_code_map``: a second cache over one file means a test that swaps
    the config has to remember to clear both, and the expensive read here
    (``_summary_rows``, the workbook) is cached already.
    """
    import yaml

    from ..domain.period import Month

    path = _repo_root() / _REF_CODE_MAP_REL
    with path.open() as f:
        cfg = yaml.safe_load(f) or {}
    declared = set((cfg.get("primes") or {}).keys())
    out: dict[str, str] = {}
    for prime, month in (cfg.get("retired_from") or {}).items():
        if str(prime) not in declared:
            # A typo here would otherwise be a silent no-op: the name matches
            # no prime, the real prime keeps earning DR, and nothing says so.
            raise ValueError(
                f"{_REF_CODE_MAP_REL}: retired_from.{prime!r} is not a prime "
                f"declared under `primes:` (have {sorted(declared)}) — as "
                "written this retirement would never apply to anything."
            )
        try:
            out[str(prime)] = str(Month.parse(str(month)))
        except Exception as exc:
            raise ValueError(
                f"{_REF_CODE_MAP_REL}: retired_from.{prime} = {month!r} is not "
                f"a 'YYYY-MM' month ({exc})"
            ) from exc
    return out


@functools.lru_cache(maxsize=1)
def _ref_code_map() -> tuple[dict[str, str], frozenset]:
    """``({ref_code: prime_id}, {deliberately-unattributed ref codes})`` from
    ``config/dr_ref_codes.yaml``.

    Raises on a duplicate code: a code owned by two primes would double-pay,
    and that is not something to degrade gracefully through.
    """
    import yaml

    path = _repo_root() / _REF_CODE_MAP_REL
    with path.open() as f:
        cfg = yaml.safe_load(f) or {}

    owner: dict[str, str] = {}
    seen_at: dict[str, str] = {}

    def _claim(code, where: str) -> str:
        c = str(code).strip()
        if c in seen_at:
            raise ValueError(
                f"{_REF_CODE_MAP_REL}: ref code {c!r} listed twice "
                f"({seen_at[c]} and {where}). Each code must appear once — "
                f"a duplicate would double-pay or mis-attribute DR."
            )
        seen_at[c] = where
        return c

    for prime_id, codes in (cfg.get("primes") or {}).items():
        for code in codes or ():
            owner[_claim(code, f"primes.{prime_id}")] = prime_id

    unattributed: set[str] = set()
    for bucket, codes in (cfg.get("unattributed") or {}).items():
        for code in codes or ():
            unattributed.add(_claim(code, f"unattributed.{bucket}"))

    return owner, frozenset(unattributed)


def dr_primes() -> frozenset:
    """Prime ids that have at least one ref code attributed to them."""
    return frozenset(_ref_code_map()[0].values())


@functools.lru_cache(maxsize=1)
def _summary_rows() -> tuple | None:
    """The DR ``Summary`` tab as a tuple of row-tuples (header first), or
    ``None`` when the workbook/sheet is unavailable. Cached: the workbook is
    read once per process (one ``--dr-only`` run reuses it across months)."""
    wb_path = _repo_root() / _WORKBOOK_REL
    if not wb_path.exists():
        log.warning(
            "DR workbook not found at %s — skipping distribution rewards "
            "(is the settle-dr-dune submodule initialised?)", wb_path,
        )
        return None

    import openpyxl  # lazy — only when DR is actually read

    wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    try:
        if _SHEET not in wb.sheetnames:
            log.warning("DR workbook %s has no %r sheet — skipping", wb_path, _SHEET)
            return None
        rows = tuple(wb[_SHEET].iter_rows(values_only=True))
    finally:
        wb.close()
    return rows or None


def _col_for(header: tuple, label: str) -> int | None:
    """Column index whose header matches ``label`` (case-insensitive). Month
    headers may be stored as text (``'YYYY-MM'``) or as Excel date cells (which
    openpyxl returns as ``datetime`` under ``data_only``); both are normalised
    to ``'YYYY-MM'`` before matching, so a workbook regeneration that switches
    the month headers to dates doesn't silently drop all DR."""
    want = label.strip().lower()
    for i, h in enumerate(header):
        if h is None:
            continue
        if hasattr(h, "year") and hasattr(h, "month"):  # date / datetime cell
            cell = f"{h.year}-{h.month:02d}"
        else:
            cell = str(h).strip()
        if cell.lower() == want:
            return i
    return None


def load_dr(prime_id: str, month: str) -> dict | None:
    """DR for ``(prime_id, month='YYYY-MM')``.

    Returns ``{"total": Decimal, "rows": [{"ref_code": str, "amount":
    Decimal, "notes": str}], "month": month}`` or ``None`` when unavailable.

    A prime retired from DR (``retired_from`` in the ref-code config) returns
    a total of ``0`` with no rows from its cutoff month on — a real zero, NOT
    ``None``, so a re-render overwrites any previously-written DR instead of
    leaving it in place.
    """
    owner, unattributed = _ref_code_map()
    if prime_id not in set(owner.values()):
        return None

    # Retired from this month on. Return an explicit ZERO rather than None:
    # None means "no data, leave the report alone", which would strand the
    # last non-zero DR on an already-written provenance when
    # ``refresh_dr_only`` runs over it. A zero total also empties
    # ``dr_breakdown``, so the per-ref-code table drops out of the summary.
    retired = _dr_retired_from().get(prime_id)
    if retired is not None and month >= retired:
        log.info(
            "DR retired for %s from %s — reporting $0 for %s",
            prime_id, retired, month,
        )
        return {"total": Decimal("0"), "rows": [], "month": month}

    rows = _summary_rows()
    if not rows:
        return None

    header = rows[0]
    col = _col_for(header, month)
    if col is None:
        log.warning(
            "DR sheet %r has no column for month %s — skipping DR for %s",
            _SHEET, month, prime_id,
        )
        return None
    ref_col = _col_for(header, "ref_code")
    if ref_col is None:
        log.warning("DR sheet %r has no 'ref_code' column — skipping", _SHEET)
        return None
    notes_col = _col_for(header, "notes")  # resolved by label, not a magic index

    out_rows: list[dict] = []
    unknown: list[tuple[str, Decimal]] = []
    for row in rows[1:]:
        ref = row[ref_col] if len(row) > ref_col else None
        if ref is None:
            continue
        code = str(ref).strip()
        if not code or code.lower() == "total":
            continue
        who = owner.get(code)
        if who is None:
            if code not in unattributed:
                unknown.append((code, _dec(row[col])))
            continue
        if who != prime_id:
            continue
        notes = row[notes_col] if notes_col is not None and len(row) > notes_col else None
        out_rows.append({
            "ref_code": code,
            "amount": _dec(row[col]),
            "notes": str(notes).strip() if notes else "",
        })

    if unknown:
        # Loud on purpose: an unmapped code is DR that reaches no prime. It is
        # either a genuine new referral needing an owner, or a deliberate
        # exclusion that belongs in `unattributed:`. Silence would let real
        # money disappear between the workbook and the settlement.
        log.error(
            "DR %s: %d ref code(s) in %s are absent from %s and were EXCLUDED "
            "— $%s total for this month: %s. Add each to a prime or to "
            "`unattributed:`.",
            month, len(unknown), _SHEET, _REF_CODE_MAP_REL,
            f"{sum((a for _, a in unknown), Decimal('0')):,.2f}",
            ", ".join(f"{c} (${a:,.2f})" for c, a in unknown),
        )

    if not out_rows:
        return None
    # The HyperSync sheet has no per-group Total row — the group total IS the
    # sum of the prime's codes.
    total = sum((r["amount"] for r in out_rows), Decimal("0"))

    return {"total": total, "rows": out_rows, "month": month}


def enrich_with_dr(pnl):
    """Populate ``distribution_rewards`` + ``dr_breakdown`` on ``pnl`` from
    the DR workbook for ``pnl``'s prime + month.

    Idempotent: sets ``distribution_rewards`` to the group total and bumps
    ``monthly_pnl`` by the *delta* vs the current value, so re-enriching an
    already-enriched pnl is a no-op (and the ``monthly_pnl`` invariant —
    which includes ``distribution_rewards`` — stays intact). Returns ``pnl``
    unchanged when no DR data is available. ``prime_agent_total_revenue`` is a
    computed property, so it updates automatically.
    """
    month = f"{pnl.month.year}-{pnl.month.month:02d}"
    dr = load_dr(pnl.prime_id, month)
    if dr is None:
        return pnl
    total = dr["total"]
    delta = total - pnl.distribution_rewards
    return dataclasses.replace(
        pnl,
        distribution_rewards=total,
        monthly_pnl=pnl.monthly_pnl + delta,
        dr_breakdown=dr["rows"],
    )
