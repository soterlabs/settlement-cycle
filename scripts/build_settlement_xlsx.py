"""Render an MSC-native settlement xlsx from a settlement output.

Self-sufficient: reads ``provenance.json`` only (plus prime YAML and SDE
config for the static reference tabs). The per-venue ``cof_alloc /
profit_to_sky / profit_to_grove`` re-attribution is computed in-process
via ``settle.load.grove_sheet.compute_sheet_rows``.

Tabs:

  1. Summary           — Prime side / Sky side / Grove-comparable Σ P2G / period info
  2. Venues            — per-venue P&L breakdown with CoF re-attribution
  3. Sky Revenue       — how sky_revenue is built (CoF + SDE) + subsidy params
  4. Sky Direct        — active Sky-Direct entries this period
  5. Debt              — per-day cum_debt (frob+grab), deductions, APYs, Sky charge
  6. Off-protocol holdings (when present)
  7. SDE daily         — per-day Sky/Grove/in-flight (when a burn occurred)

Inputs:
  settlements/{prime}/{month}/provenance.json    (canonical machine-readable output)
  config/{prime}.yaml
  config/sky_direct_exposures.yaml
  config/subsidy_reference_rates.yaml            (for the ref-rate readout)

Output:
  settlements/{prime}/{month}/{prime}_settlement_{month_name}_{year}.xlsx
  e.g. grove_settlement_april_2026.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parent.parent

# Import the shared grove-sheet computation. The path manipulation lets the
# script run as a CLI without installing the package.
sys.path.insert(0, str(_REPO / "src"))
from settle.load.grove_sheet import compute_sheet_rows  # noqa: E402
from settle.load.summary import _venue_sort_key  # noqa: E402

# Styling.
_BOLD   = Font(bold=True)
_TITLE  = Font(bold=True, size=14)
_MUTED  = Font(color="666666", italic=True)
_HEADER_FILL = PatternFill("solid", fgColor="DDE6F1")
_SUBTLE_FILL = PatternFill("solid", fgColor="F4F7FB")
_USD     = '"$"#,##0.00;"−$"#,##0.00;"$"0.00'
_USD0    = '"$"#,##0;"−$"#,##0;"$"0'
_PCT     = '0.000%'
_THIN    = Side(style="thin", color="C5C9CC")
_BOX     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _D(x) -> Decimal:
    if x is None or x == "":
        return Decimal("0")
    if isinstance(x, Decimal):
        return x
    return Decimal(str(x))


def _set_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL


# --------------------------------------------------------------------------
# IO
# --------------------------------------------------------------------------

def _read_provenance(cell: Path) -> dict:
    with (cell / "provenance.json").open(encoding="utf-8") as f:
        return json.load(f)


def _sheet_rows_as_strings(sheet_rows: list[dict]) -> list[dict]:
    """Convert ``compute_sheet_rows`` output (per-venue Decimal values) into
    string-keyed dicts to match the legacy CSV-row shape that the xlsx
    writers historically consumed.

    The writers (``_write_summary``, ``_write_venues``, ``_write_sky_revenue``,
    ``_write_sde``) read fields via ``r["foo"]`` then ``_D(r["foo"])`` —
    accepting either CSV strings or Decimals already. We pass Decimals
    straight through; the writers' ``_D`` parsing is a no-op for them.
    """
    return sheet_rows


def _read_prime_yaml(prime_id: str) -> dict:
    with (_REPO / "config" / f"{prime_id}.yaml").open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def _read_sde(prime_id: str, period_start: date) -> list[dict]:
    with (_REPO / "config" / "sky_direct_exposures.yaml").open(encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    out: list[dict] = []
    for section in ("active", "historical"):
        for e in cfg.get(section, []) or []:
            if e.get("prime") != prime_id:
                continue
            if e.get("kind") == "pattern":
                # Pattern entries apply at the prime/PSM3 level — not per-venue.
                # Skipping here keeps the SDE tab focused on venue-level entries.
                continue
            start = e["start_date"]
            end = e.get("end_date")
            if isinstance(start, str): start = date.fromisoformat(start)
            if isinstance(end, str):   end   = date.fromisoformat(end)
            active = start <= period_start and (end is None or end >= period_start)
            out.append({**e, "_start": start, "_end": end, "_active": active})
    return out


# --------------------------------------------------------------------------
# Tab writers
# --------------------------------------------------------------------------

def _write_summary(ws, prov: dict, sheet_rows: list[dict]) -> None:
    ws.title = "Summary"
    prime = prov["prime_id"].upper()
    month = prov["month"]
    res   = prov["results"]
    par   = _D(res["prime_agent_revenue"])
    ar    = _D(res["agent_rate"])
    dr    = _D(res.get("distribution_rewards") or 0)
    par_t = _D(res.get("prime_agent_total_revenue", par + ar + dr))
    sky   = _D(res["sky_revenue"])
    sd    = sum((_D(r["sd_revenue"]) for r in sheet_rows), Decimal("0"))
    cof   = sky - sd
    sum_p2g = sum((_D(r["profit_to_grove"]) for r in sheet_rows), Decimal("0"))
    monthly_pnl = par + ar + dr - sky
    # New display-only surfaces from PR #104 — see provenance schema.
    sky_gross    = _D(res.get("sky_revenue_gross") or 0)
    spread_reimb = _D(res.get("susds_spread_reimbursement") or 0)

    ws.append([f"{prime} — Monthly settlement {month}"])
    ws["A1"].font = _TITLE
    ws.append([])

    def _block(title: str, rows: list[tuple[str, Decimal]], total: Decimal) -> None:
        """Three-block layout: header → addends → blank-label total row."""
        ws.append([title, "USD"])
        _header_row(ws, ws.max_row, 2)
        for lbl, val in rows:
            ws.append([lbl, float(val)])
            ws.cell(ws.max_row, 2).number_format = _USD
        # Total row — no label (matches the cleaned reference layout)
        ws.append(["", float(total)])
        cell = ws.cell(ws.max_row, 2)
        cell.number_format = _USD
        cell.font = _BOLD
        cell.border = Border(top=_THIN)

    prime_rows = [
        ("prime_agent_revenue (gross venue yield to prime)", par),
        ("+ agent_rate (subproxy USDS / sUSDS yield)",       ar),
    ]
    if dr != 0:
        prime_rows.append(("+ distribution_rewards (settle-dr-dune)", dr))
    # Chronicle Points (Grove only) — keeps the tab's addends summing to
    # prime_agent_total_revenue, which includes the component.
    cp = _D(res.get("chronicle_points") or 0)
    if cp != 0:
        prime_rows.append(("+ chronicle_points (20% of base rate on Chronicle Farm USDS)", cp))
    # Governance Accessibility Rewards (Skybase only) — same convention.
    # The rate lives in config (gar.share); the provenance gar_basis string
    # carries the actual derivation, so don't hardcode a percentage here.
    gar = _D(res.get("gar") or 0)
    if gar != 0:
        gar_note = str(res.get("gar_basis") or "share × prior-month Sky Net Revenue")
        prime_rows.append((f"+ governance_accessibility_rewards ({gar_note})", gar))
    _block("Prime side", rows=prime_rows, total=par_t)
    ws.append([])

    _block(
        "Sky side",
        rows=[
            ("CoF on utilized (BR × Net_Subs)",          cof),
            ("+ SDE revenue (Sky-Direct, full to Sky)",  sd),
        ],
        total=sky,
    )
    ws.append([])

    # Sky Revenue (max) — display-only ceiling on BR (no idle/SDE deductions).
    # NOT a true ceiling on sky_revenue: actual sky_revenue adds sde_revenue
    # on top of BR-on-utilized, so for primes with material SDE positions
    # sky_revenue can exceed sky_revenue_gross. The subsidised BR (ref_rate
    # ramp) is already applied — the rate matches actual sky_revenue, not
    # the raw Maker base rate.
    if sky_gross > 0:
        _block(
            "Sky Revenue (max) — BR × full ilk debt, no deductions",
            rows=[
                ("CoF on Net_Subs (actual BR × utilized)",         cof),
                ("reduction from idle/SDE deductions",             -(sky_gross - cof)),
            ],
            total=sky_gross,
        )
        ws.append([])

    # sUSDS spread — deducted from sky_revenue (Sky charges full BR on the
    # underlying utilized then refunds 30 bps; net cost to prime = SSR × V).
    # Surfaced here for audit visibility.
    if spread_reimb != 0:
        ws.append(["sUSDS spread (Curve LP + PSM3) — deducted from sky_revenue", float(spread_reimb)])
        ws.cell(ws.max_row, 2).number_format = _USD
        ws.cell(ws.max_row, 1).font = _MUTED
        ws.append([])

    _block(
        'Comparison (Grove-style "Profit to Grove")',
        rows=[
            ("prime_agent_revenue",                    par),
            ("− CoF (deducted per-venue in display)",  -cof),
        ],
        total=sum_p2g,
    )
    ws.append([])

    # Period info
    ws.append(["Period",     f"{prov['period']['start']} → {prov['period']['end']} "
                              f"({prov['period']['n_days']} days)"])
    ws.append(["Pin blocks", ", ".join(f"{k}={v}" for k, v in (prov.get('pin_blocks_eom') or {}).items())])
    ws.append(["Generated",  prov.get("generated_at_utc", "")])
    ws.append(["Pipeline",   prov.get("settle_version", "")])

    _set_widths(ws, {1: 60, 2: 22})


def _write_venues(ws, sheet_rows: list[dict], prime_cfg: dict) -> None:
    ws.title = "Venues"
    # Index venues from prime config to enrich with chain + category.
    by_id = {v.get("id"): v for v in prime_cfg.get("venues", []) or []}

    cols = [
        "Venue", "Label", "Chain", "Pricing cat.",
        "Position SoM", "Position EoM", "Period inflow",
        "actual_revenue", "external_revenue", "revenue (to prime)",
        "sd_revenue (to Sky)",
        "Avg value", "Weight", "CoF alloc", "Profit to Sky", "Profit to Grove",
        "Utilized Deduction (avg)", "Spread Reimb",
        "Notes",
    ]
    ws.append(cols)
    _header_row(ws, 1, len(cols))

    # Venues with ``hide_per_venue_pnl`` are not rendered in the PnL body;
    # they appear in a position-only sub-table below. The venue still
    # contributes to ``prime_agent_revenue`` at the prime level (the
    # Summary tab's headline is unchanged) — only the per-venue PnL row
    # is suppressed. Currently set on Spark Savings V2 vaults
    # (S56/S57/S59/S60) whose per-venue "revenue" is the negative
    # VSR-liability accrual on depositor capital.
    pnl_rows      = [r for r in sheet_rows if not r.get("hide_per_venue_pnl")]
    position_only = [r for r in sheet_rows if     r.get("hide_per_venue_pnl")]

    # Sort by absolute Profit to Sky desc so Grove-large positions surface first.
    for r in sorted(pnl_rows, key=lambda x: abs(float(x["profit_to_sky"])), reverse=True):
        vid = r["venue_id"]
        v = by_id.get(vid, {})
        ws.append([
            vid,
            r["label"],
            v.get("chain", ""),
            v.get("pricing_category", ""),
            float(r["value_som"]),
            float(r["value_eom"]),
            float(_D(r["value_eom"]) - _D(r["value_som"])),   # period_inflow proxy
            float(r["actual_rev"]),
            float(r["external"]),
            float(r["revenue"]),
            float(r["sd_revenue"]),
            float(r["avg_value"]),
            float(r["weight"]),
            float(r["cof_alloc"]),
            float(r["profit_to_sky"]),
            float(r["profit_to_grove"]),
            float(_D(r.get("deduction_avg") or 0)),
            float(_D(r.get("spread_reimb")  or 0)),
            r.get("note", ""),
        ])

    body_last_row = ws.max_row

    # ── Position-only sub-section ──────────────────────────────────
    # PnL-suppressed venues rendered with positions only. Their
    # ``actual_revenue`` (negative VSR liability for Savings V2) flows
    # into the prime-level ``prime_agent_revenue`` headline on the
    # Summary tab; per-venue PnL is intentionally hidden.
    if position_only:
        ws.append([])
        ws.append(["Position-only venues (PnL aggregated at prime level)"])
        ws.cell(ws.max_row, 1).font = _BOLD
        ws.append([
            "Venue", "Label", "Chain", "Pricing cat.",
            "Position SoM", "Position EoM",
        ])
        _header_row(ws, ws.max_row, 6)
        # Sort with the same numeric-aware key as ``summary.py`` so the
        # two surfaces stay in step (e.g. S9 before S10).
        for r in sorted(position_only, key=lambda x: _venue_sort_key(x["venue_id"])):
            vid = r["venue_id"]
            v = by_id.get(vid, {})
            ws.append([
                vid,
                r["label"],
                v.get("chain", ""),
                v.get("pricing_category", ""),
                float(r["value_som"]),
                float(r["value_eom"]),
            ])
            for c in (5, 6):
                ws.cell(ws.max_row, c).number_format = _USD0
        # Aggregate cell — keeps the visible Venues body reconcilable
        # against the Summary tab's ``Profit to Grove`` total. Without
        # this, an auditor summing the visible body's ``revenue (to
        # prime)`` / ``Profit to Grove`` columns would find a gap equal
        # to the suppressed VSR liability with no in-sheet explanation.
        agg_actual = sum(
            (_D(r["actual_rev"]) for r in position_only), Decimal("0"),
        )
        ws.append([
            "",
            "Aggregated actual_revenue (included in prime_agent_revenue, "
            "not in Venues body above)",
            "", "", "",
            float(agg_actual),
        ])
        ws.cell(ws.max_row, 6).number_format = _USD0
        ws.cell(ws.max_row, 6).font = _BOLD
        ws.cell(ws.max_row, 2).font = _MUTED

    # Number formats for the PnL body: USD cols are 5–12, 14–18. Pct col
    # is 13. Last col is text. Limit the formatting loop to the PnL body
    # so the position-only sub-section above keeps its own narrower
    # formatting.
    for row in range(2, body_last_row + 1):
        for c in (5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18):
            ws.cell(row, c).number_format = _USD0
        ws.cell(row, 13).number_format = _PCT

    _set_widths(ws, {
        1: 7, 2: 50, 3: 11, 4: 13,
        5: 16, 6: 16, 7: 16, 8: 14, 9: 14, 10: 16, 11: 16,
        12: 16, 13: 9, 14: 14, 15: 16, 16: 16,
        17: 22, 18: 14, 19: 30,
    })
    ws.freeze_panes = "C2"


def _write_subsidy_panel(ws, summary: dict) -> None:
    """Rates & subsidy headline panel — FORMAT ONLY.

    Renders the per-period aggregates that ``compute.sky_revenue.
    summarize_subsidy`` emits into ``provenance.json["subsidy_summary"]``.
    All economics (tranche split, effective rate, $ benefit, zero-benefit
    flag) are computed in the compute layer against the exact rate schedule
    charged, so this panel does no math and cannot drift from settlement.
    """
    cap = _D(summary["cap_usd"])
    kind = summary.get("ref_rate_kind", "ref rate")

    def _row(label, val, fmt=None, bold=False):
        ws.append([label, val])
        if fmt is not None:
            ws.cell(ws.max_row, 2).number_format = fmt
        if bold:
            ws.cell(ws.max_row, 1).font = _BOLD
            ws.cell(ws.max_row, 2).font = _BOLD

    def _pct(key):  # rate may be None (no subsidy-active days)
        return (float(summary[key]), _PCT) if summary.get(key) is not None else ("n/a", None)

    def _pct_compat(apr_key, apy_key):
        """Same, tolerating the pre-2026-09-01 ``*_apy`` key names."""
        v = _rate(summary, apr_key, apy_key)
        return (float(v), _PCT) if v is not None else ("n/a", None)

    ws.append(["Rates & subsidy — effective rate charged this period"])
    ws.cell(ws.max_row, 1).font = _BOLD
    ws.append(["Metric", "Value"]); _header_row(ws, ws.max_row, 2)
    _row("Time-weighted utilized", float(_D(summary["tw_utilized"])), _USD0)
    _row(f"  — first ${cap/Decimal(10**9):.0f}B (subsidised tranche)",
         float(_D(summary["sub_tranche_balance"])), _USD0)
    _row("  — excess (full base-rate tranche)",
         float(_D(summary["exc_tranche_balance"])), _USD0)
    _row("Base rate (BR_apr = apy_to_apr(SSR,12) + spread; 30bps, 20bps from 2026-07-23), period avg", float(_rate(summary, "base_apr_avg", "base_apy_avg")), _PCT)
    _row(f"Reference rate ({kind}), period avg", *_pct_compat("ref_apr_avg", "ref_apy_avg"))
    _row("Subsidised rate (BR*), period avg", *_pct_compat("sub_apr_avg", "sub_apy_avg"))
    _row("Effective blended rate charged", float(_rate(summary, "effective_apr", "effective_apy")), _PCT, bold=True)
    _row("Diff vs base rate (bps)", float(summary["diff_bps"]), "0.0", bold=True)
    _row("Subsidy benefit to prime (USD)", float(_D(summary["subsidy_benefit"])), _USD, bold=True)

    # Subsidy dollar reconciliation (full BR − actual = benefit).
    ws.append([])
    _row("CoF at full base rate (no subsidy)", float(_D(summary["full_br_cof"])), _USD)
    _row("  − actual CoF (subsidy on first tranche)", float(_D(summary["actual_cof"])), _USD)
    _row("  = subsidy benefit", float(_D(summary["subsidy_benefit"])), _USD, bold=True)

    # Tranche split — balance + CoF (sub_tranche_cof + exc_tranche_cof =
    # actual_cof exactly, both from compute).
    ws.append([])
    ws.append(["Tranche", "Avg balance", "CoF (USD)"])
    _header_row(ws, ws.max_row, 3)
    for lbl, bal_key, cof_key in [
        (f"Subsidised (first ${cap/Decimal(10**9):.0f}B)", "sub_tranche_balance", "sub_tranche_cof"),
        ("Excess (> cap)", "exc_tranche_balance", "exc_tranche_cof"),
    ]:
        ws.append([lbl, float(_D(summary[bal_key])), float(_D(summary[cof_key]))])
        ws.cell(ws.max_row, 2).number_format = _USD0
        ws.cell(ws.max_row, 3).number_format = _USD

    # ⚠️ zero-benefit flag — same condition as the compute-layer warning
    # (no day had sub_apr < base_apr), so panel and log never disagree.
    if summary.get("zero_benefit"):
        ws.append([
            "⚠️ Subsidy produced ≈$0 benefit this period — the reference "
            f"rate ({kind}) sits at/above the base rate every day. Verify it "
            "is current; a stale or placeholder value silently nullifies the "
            "subsidy (this caused the May 2026 Spark mis-pricing)."
        ])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="B45309")
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 45
    ws.append([])


def _write_sky_revenue(ws, prov: dict, sheet_rows: list[dict], prime_cfg: dict) -> None:
    ws.title = "Sky Revenue"
    res = prov["results"]
    sky = _D(res["sky_revenue"])
    sd  = sum((_D(r["sd_revenue"]) for r in sheet_rows), Decimal("0"))
    cof = sky - sd
    sky_gross    = _D(res.get("sky_revenue_gross") or 0)
    spread_reimb = _D(res.get("susds_spread_reimbursement") or 0)

    ws.append(["How Sky's monthly take is built"])
    ws["A1"].font = _TITLE
    ws.append([])

    # Rates & subsidy headline panel — only when compute emitted the
    # per-period aggregates (subsidised-borrowing primes: Spark, Grove).
    # The panel formats this dict; it does no economics of its own.
    if prov.get("subsidy_summary"):
        _write_subsidy_panel(ws, prov["subsidy_summary"])

    ws.append(["Component", "USD"])
    _header_row(ws, ws.max_row, 2)
    rows = [
        ("CoF on utilized debt (= Σ_d max(utilized_d,0) × (1+BR_d)^(1/365)−1)", cof),
        ("+ Sky-Direct revenue (full venue yield on fixed/capped SDE)",         sd),
        ("equals sky_revenue (total Sky take this month)",                       sky),
    ]
    for lbl, val in rows:
        ws.append([lbl, float(val)])
        ws.cell(ws.max_row, 2).number_format = _USD

    ws.append([])
    ws.append(["Utilized base = cum_debt − idle USDS − PSM USDS − Σ SDE asset value − Curve idle − lending idle"])
    ws.cell(ws.max_row, 1).font = _MUTED
    ws.append([])

    # Sky Revenue (max) reconciliation — pure BR × cum_debt vs actual CoF.
    # Display-only diagnostic that surfaces how much the idle/SDE/PSM/Curve
    # deductions reduce the BR component of sky_revenue.
    if sky_gross > 0:
        ws.append(["Sky Revenue (max) — BR × full ilk debt (no deductions)", float(sky_gross)])
        ws.cell(ws.max_row, 2).number_format = _USD
        ws.cell(ws.max_row, 1).font = _BOLD
        ws.append(["    ↳ actual CoF on Net_Subs (BR × utilized)",            float(cof)])
        ws.cell(ws.max_row, 2).number_format = _USD
        ws.append(["    ↳ reduction from idle/SDE deductions",                -float(sky_gross - cof)])
        ws.cell(ws.max_row, 2).number_format = _USD
        ws.append([
            "Note: sky_revenue can EXCEED sky_revenue_gross for primes with "
            "active SDE positions — sky_revenue also adds sde_revenue on top "
            "of BR-on-utilized. The subsidised BR (ref_rate ramp) is already "
            "applied here; this is NOT the raw Maker base rate."
        ])
        ws.cell(ws.max_row, 1).font = _MUTED
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 60
        ws.append([])

    # sUSDS spread — Sky Revenue reduction (Cat B refund).
    if spread_reimb != 0:
        ws.append([
            "sUSDS spread (Curve LP + PSM3) — the BR−SSR spread (30 bps; "
            "20 bps from 2026-07-23) × value, integrated daily, deducted "
            "from sky_revenue. Sky charges full BR on the underlying "
            "utilized; this row is the offsetting refund to the prime so "
            "SSR + BR + spread nets to zero economically.",
            float(spread_reimb),
        ])
        ws.cell(ws.max_row, 2).number_format = _USD
        ws.cell(ws.max_row, 1).font = _MUTED
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 60
        ws.append([])

    # Subsidy params if enabled.
    sub_cfg = prime_cfg.get("subsidy") or {}
    if sub_cfg.get("enabled"):
        ws.append(["Subsidy", "Value"])
        _header_row(ws, ws.max_row, 2)
        ws.append(["enabled",        "true"])
        ws.append(["cap_usd",        sub_cfg.get("cap_usd", "")])
        ws.cell(ws.max_row, 2).number_format = _USD0
        ws.append(["program_start",  sub_cfg.get("program_start", "")])
        ws.append(["ramp_months",    sub_cfg.get("ramp_months", "")])
        # ref_rate_kind may be a scalar kind or a dated list of
        # {kind, from} entries (the 2026-07-23 tbill_3m→sofr switch) —
        # render the list as the same compact label SubsidyConfig builds.
        raw_kind = sub_cfg.get("ref_rate_kind", "")
        if isinstance(raw_kind, list):
            raw_kind = raw_kind[0]["kind"] + "".join(
                f"→{e['kind']}@{e['from']}" for e in raw_kind[1:]
            )
        ws.append(["ref_rate_kind",  raw_kind])

        # T value for this period
        from settle.domain.subsidy import months_elapsed_since
        pstart = sub_cfg.get("program_start")
        if isinstance(pstart, str):
            pstart = date.fromisoformat(pstart)
        period_start = date.fromisoformat(prov["period"]["start"])
        period_end   = date.fromisoformat(prov["period"]["end"])
        t_start = months_elapsed_since(period_start, pstart)
        t_end   = months_elapsed_since(period_end,   pstart)
        ws.append(["T this period",  f"{t_start} (SoM) → {t_end} (EoM)"])
        ws.append([
            "Formula",
            "subsidised_apr = ref_rate + (base_apr − ref_rate) × min(T, 24) / 24, "
            "applied to first cap_usd of utilized; excess at full base_apr",
        ])
        ws.cell(ws.max_row, 1).font = _MUTED

    ws.append([])
    ws.append(["Base rate composition: base_apr = apy_to_apr(SSR, n=12) + spread (nominal; spread 30bps, 20bps from 2026-07-23). Daily charge = utilized_d x base_apr / 365 — no intra-month compounding."])
    ws.cell(ws.max_row, 1).font = _MUTED

    _set_widths(ws, {1: 80, 2: 22})


def _write_sde(ws, sde: list[dict], sheet_rows: list[dict]) -> None:
    ws.title = "Sky Direct"
    ws.append(["Sky-Direct Exposures applied this period"])
    ws["A1"].font = _TITLE
    ws.append([])

    cols = ["Venue", "Kind", "Cap (USD)", "Start", "End", "Active?", "Source", "Label",
            "actual_revenue", "sd_revenue (to Sky)"]
    ws.append(cols)
    _header_row(ws, ws.max_row, len(cols))

    by_id = {r["venue_id"]: r for r in sheet_rows}
    for e in sde:
        vid = e.get("venue_id", "")
        row = by_id.get(vid, {})
        ws.append([
            vid,
            e.get("kind", ""),
            e.get("cap_usd", "") or "",
            str(e["_start"]) if e.get("_start") else "",
            str(e["_end"]) if e.get("_end") else "—",
            "YES" if e["_active"] else "no",
            e.get("source", ""),
            e.get("label", ""),
            float(_D(row.get("actual_rev", 0))) if row else 0,
            float(_D(row.get("sd_revenue", 0))) if row else 0,
        ])
        if isinstance(e.get("cap_usd"), (int, float)):
            ws.cell(ws.max_row, 3).number_format = _USD0
        ws.cell(ws.max_row, 9).number_format  = _USD0
        ws.cell(ws.max_row, 10).number_format = _USD0

    _set_widths(ws, {1: 8, 2: 8, 3: 14, 4: 12, 5: 12, 6: 8, 7: 50, 8: 25, 9: 16, 10: 18})


def _write_off_protocol_holdings(ws, prov: dict, prime_cfg: dict) -> None:
    """Off-protocol holdings — display-only venues. Surfaced for visibility
    but excluded from prime_agent_revenue, sky_revenue, and the NAV cost-
    basis invariant. Realized P&L on any round-trip lands at the anchor
    venue when the cash arrives at the ALM — see
    ``_cat_a_capital_inflow_timeseries`` paired-principal-cap classifier.
    """
    ws.title = "Off-protocol holdings"
    rows = prov.get("display_only_breakdown") or []
    ws.append(["Off-protocol holdings (tracked, not in prime revenue)"])
    ws["A1"].font = _TITLE
    ws.append([])
    ws.append([
        "These positions sit with external counterparties (e.g. OOB OTC "
        "venues, off-protocol custodians). The balances below are visible "
        "for monthly-NAV reporting only — they do NOT contribute to "
        "prime_agent_revenue or sky_revenue. Any realized gain or loss on "
        "the round-trip is booked at the anchor venue when the cash "
        "settles back at the ALM proxy."
    ])
    ws.append([])

    if not rows:
        ws.append(["(no display-only venues active this period)"])
        _set_widths(ws, {1: 90})
        return

    by_id = {v.get("id"): v for v in prime_cfg.get("venues", []) or []}

    cols = [
        "Venue", "Label", "Chain", "Counterparty (paired_source)",
        "Anchor venue (paired_with)",
        "Outstanding SoM (USD)", "Outstanding EoM (USD)", "Δ over period",
        "Status",
    ]
    ws.append(cols)
    _header_row(ws, ws.max_row, len(cols))

    for r in rows:
        vid = r["venue_id"]
        cfg = by_id.get(vid, {})
        chain = cfg.get("chain", "")
        paired_source = (cfg.get("paired_source") or "")
        paired_with = (cfg.get("paired_with") or "")
        som = _D(r["value_som"])
        eom = _D(r["value_eom"])
        delta = eom - som
        if eom == 0 and som > 0:
            status = "fully returned"
        elif eom == som and som > 0:
            status = "outstanding (no change)"
        elif eom < som:
            status = "partial return"
        elif eom > som:
            status = "new principal-out"
        else:
            status = "—"
        ws.append([
            vid, r.get("label", ""), chain, paired_source, paired_with,
            float(som), float(eom), float(delta), status,
        ])
        for col in (6, 7, 8):
            ws.cell(ws.max_row, col).number_format = _USD0

    ws.append([])
    ws.append([
        "Note: realized gain on returns (the round-trip spread) is "
        "recognized as revenue at the anchor venue via the paired-"
        "principal-cap classifier — see provenance.json venue_breakdown "
        "for the anchor's actual_revenue / external_revenue split."
    ])

    _set_widths(ws, {
        1: 8, 2: 50, 3: 12, 4: 46, 5: 20,
        6: 18, 7: 18, 8: 16, 9: 22,
    })


def _write_sde_daily(ws, prov: dict) -> None:
    """Per-day Sky / Grove / In-flight NAV / Total breakdown for capped
    SDE venues with an on-chain burn event (``burn_date`` set).

    Phase rule per day:
      * day < burn_date            → Sky = cum_value (cap), Grove = uncapped − cum_value, in-flight = 0
      * burn_date ≤ day ≤ in_flight_end → Sky = 0, Grove = uncapped (residual), in-flight = cum_value
      * day > in_flight_end        → Sky = 0, Grove = uncapped, in-flight = 0

    where ``in_flight_end = usdc_settlement_date or end_date``. The total
    column = sum of the three (= the prime+Sky combined NAV attributable
    to this venue on that day).
    """
    breakdown = prov.get("sde_daily_breakdown") or []
    venues_with_burn = [b for b in breakdown if b.get("burn_date")]
    if not venues_with_burn:
        # No burn events this period — skip the tab entirely.
        return
    ws.title = "SDE daily"
    ws.append(["SDE per-day breakdown (capped venues with on-chain burn)"])
    ws["A1"].font = _TITLE
    ws.append([])
    ws.append([
        "For each capped SDE venue with an on-chain burn, the daily on-"
        "chain position is decomposed into Sky's slice (cap-protected, "
        "pre-burn), Grove's residual (above the cap, surviving the burn), "
        "and the in-flight USDC settling back to the ALM (post-burn / "
        "pre-settlement). The three columns sum to the total NAV "
        "attributable to this venue on that day."
    ])
    ws.append([])

    for b in venues_with_burn:
        cap_str = b.get("cap_usd") or ""
        burn = date.fromisoformat(b["burn_date"])
        in_flight_end_raw = b.get("usdc_settlement_date") or b.get("end_date")
        in_flight_end = (
            date.fromisoformat(in_flight_end_raw) if in_flight_end_raw else None
        )
        end_date_str = b.get("end_date") or "—"
        settle_str = b.get("usdc_settlement_date") or "—"

        ws.append([f"{b['venue_id']} — {b.get('label', '')}"])
        ws.cell(ws.max_row, 1).font = _BOLD
        ws.append([
            f"Cap: ${float(_D(cap_str)):,.0f}" if cap_str else "Cap: —",
            f"burn_date: {b['burn_date']}",
            f"usdc_settlement_date: {settle_str}",
            f"end_date: {end_date_str}",
        ])
        ws.append([])
        ws.append(["Date", "Sky position", "Grove position", "In-flight NAV", "Total"])
        _header_row(ws, ws.max_row, 5)

        for r in b.get("daily", []):
            d = date.fromisoformat(r["block_date"])
            cum = _D(r["cum_value"])
            unc = _D(r["uncapped_value"])
            if d < burn:
                # Pre-burn: Sky and Grove both visible on-chain. Sky's
                # share is the cap-protected slice; Grove's is the excess.
                sky = cum
                grove = max(Decimal("0"), unc - cum)
                inflight = Decimal("0")
            elif in_flight_end is not None and d <= in_flight_end and cum > 0:
                # In-flight: Sky's shares are destroyed (so uncapped only
                # reflects Grove's residual). Cap-preserved cum_value is
                # the in-flight USDC pending settlement.
                sky = Decimal("0")
                grove = unc
                inflight = cum
            else:
                # Post-settlement / post-end: Sky's slice is recognised
                # at $0 (the USDC has landed at the ALM and joined idle
                # alm_usds for separate accounting). Grove's residual
                # remains on-chain.
                sky = Decimal("0")
                grove = unc
                inflight = Decimal("0")
            total = sky + grove + inflight
            ws.append([
                d.isoformat(), float(sky), float(grove), float(inflight), float(total),
            ])
            for c in (2, 3, 4, 5):
                ws.cell(ws.max_row, c).number_format = _USD0

        ws.append([])  # blank row between venues

    _set_widths(ws, {1: 12, 2: 18, 3: 18, 4: 18, 5: 18})


def _rate(d: dict, apr_key: str, apy_key: str):
    """Read a rate that was renamed ``*_apy`` -> ``*_apr`` on 2026-09-01.

    Months settled before that date keep the old keys in their (gitignored,
    on-disk) provenance and are deliberately NOT restated, so re-rendering
    their workbooks has to keep working. Falls back to the legacy name.
    """
    v = d.get(apr_key)
    return d.get(apy_key) if v is None else v


def _write_debt(ws, prov: dict) -> None:
    """Per-day ilk-debt + Sky-charge breakdown.

    Debug-oriented tab for the prime team to reconcile two methodology
    knobs that drive ``sky_revenue``:

      * **Debt source**: ``cum_debt`` here is built from BOTH ``vat.frob``
        (regular draws/repays) and ``vat.grab`` (stability-fee
        capitalisation) trace events — matching the canonical Vat
        ``urns[ilk][u].art``. Primes using a frob-only "Subscriptions"
        spreadsheet will see lower numbers and a correspondingly lower
        CoF charge — the daily delta is the cumulative-grab dart through
        that date.

      * **Rate composition** (from 2026-09-01): ``base_apr`` =
        ``apy_to_apr(SSR, n=12) + spread``. SSR is quoted as an APY — it
        compounds per-second into the sUSDS index on-chain — while the
        spread is a governance-set APR, so SSR is converted to its
        nominal equivalent, ``12 × [(1+SSR)^(1/12) − 1]``, before the two
        are added. At SSR 3.52% + 20bps that is 3.464456% + 0.20% =
        **3.664456%**. n = 12 matches the settlement cadence, so the
        conversion round-trips: ``(1 + SSR_apr/12)^12 - 1`` returns the
        3.52% SSR APY exactly. (That identity is about the converted SSR
        leg, not ``base_apr``.)
        ``sub_apr`` is the subsidised rate after the ramp: ``ref_rate +
        (base − ref_rate) × T/24``, clamped at base_apr when the
        reference exceeds it.

      * **The charge is NOMINAL — no intra-month compounding.** Each day
        is ``utilized_d × rate_d / 365`` and the days are summed, so any
        row reconciles on its own. The compounding that does occur is the
        MSC capitalising the month's charge into the ilk debt
        (``vat.grab`` with positive ``dart``), which shows up as a step in
        ``cum_debt`` rather than inside this column.
        ``daily_sky_rev`` applies sub_apr to the first $1B of utilized and
        full base_apr to the excess; ``daily_sky_rev_gross`` applies the
        same schedule on the full cum_debt (no deductions), making the gap
        to actual a measure of the deduction stack.
    """
    rows = prov.get("sky_revenue_daily") or []
    if not rows:
        # No daily series captured (e.g. legacy provenance.json from before
        # this field was added) — skip the tab cleanly rather than render
        # an empty grid.
        return
    ws.title = "Debt"
    ws.append(["Daily ilk debt + Sky charge breakdown"])
    ws["A1"].font = _TITLE
    ws.append([])
    ws.append([
        "cum_debt = Σ on-chain Vat dart from frob (0x76088703) + grab "
        "(0x7bab3f40), then scaled by Vat.ilks[ilk].rate_d / 1e27 read "
        "at each day's EoD block — actual outstanding USDS per day, not "
        "raw normalised Art. utilized = cum_debt − Σ deductions. base_apr "
        "= apy_to_apr(SSR, n=12) + spread (nominal; spread 30bps, 20bps from "
        "2026-07-23). sub_apr applies on the "
        "first cap_usd of utilized when the subsidy is active; excess "
        "pays base_apr."
    ])
    ws.cell(ws.max_row, 1).font = _MUTED
    ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[ws.max_row].height = 45
    ws.append([])

    # If no row has subsidy data populated, omit the three subsidy columns
    # (T, ref_rate, sub_apr) — keeps the tab tight for non-subsidy primes.
    has_subsidy = any(_rate(r, "sub_apr", "sub_apy") is not None for r in rows)

    cols = [
        "Date",
        "cum_debt",
        "− ALM idle",
        "− PSM USDS",
        "− SDE NAV",
        "− Curve idle",
        "− Lending idle",
        "= utilized",
        "SSR APY",
        "base APR",
    ]
    if has_subsidy:
        cols += ["T (months)", "ref_rate APR", "sub APR"]
    cols += ["daily Sky charge", "daily Sky charge (gross on cum_debt)"]
    ws.append(cols)
    _header_row(ws, ws.max_row, len(cols))

    sum_rev   = Decimal("0")
    sum_gross = Decimal("0")
    for r in rows:
        rev = _D(r["daily_sky_rev"])
        gross = _D(r["daily_sky_rev_gross"])
        sum_rev   += rev
        sum_gross += gross
        out = [
            r["date"],
            float(_D(r["cum_debt"])),
            float(_D(r["alm_usds"])),
            float(_D(r["psm_usds"])),
            float(_D(r["sde_av"])),
            float(_D(r["curve_idle"])),
            float(_D(r["lending_idle"])),
            float(_D(r["utilized"])),
            r["ssr_apy"],
            _rate(r, "base_apr", "base_apy"),
        ]
        if has_subsidy:
            out += [r.get("t_months"), _rate(r, "ref_rate_apr", "ref_rate_apy"),
                    _rate(r, "sub_apr", "sub_apy")]
        out += [float(rev), float(gross)]
        ws.append(out)
        row_n = ws.max_row
        # USD columns: cum_debt … utilized, daily charges
        for c in (2, 3, 4, 5, 6, 7, 8):
            ws.cell(row_n, c).number_format = _USD0
        # APY columns
        for c in (9, 10):
            ws.cell(row_n, c).number_format = _PCT
        if has_subsidy:
            for c in (12, 13):  # ref_rate APR, sub APR (T stays integer)
                ws.cell(row_n, c).number_format = _PCT
        for c in (len(cols) - 1, len(cols)):
            ws.cell(row_n, c).number_format = _USD

    # Totals footer — only the additive columns (daily charges) sum
    # meaningfully across the period; the rest are point-in-time.
    ws.append([])
    total_row = [""] * len(cols)
    total_row[0] = "Σ daily charges"
    total_row[-2] = float(sum_rev)
    total_row[-1] = float(sum_gross)
    ws.append(total_row)
    row_n = ws.max_row
    ws.cell(row_n, 1).font = _BOLD
    for c in (len(cols) - 1, len(cols)):
        ws.cell(row_n, c).number_format = _USD
        ws.cell(row_n, c).font = _BOLD

    # Widths
    widths = {1: 12, 2: 16, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 16, 9: 10, 10: 10}
    if has_subsidy:
        widths.update({11: 11, 12: 12, 13: 11, 14: 18, 15: 22})
    else:
        widths.update({11: 18, 12: 22})
    _set_widths(ws, widths)


# --------------------------------------------------------------------------
# Entrypoint
# --------------------------------------------------------------------------

_MONTH_NAMES = (
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
)


def _output_filename(prime_id: str, month: str) -> str:
    """Return ``{prime}_settlement_{month_name}_{year}.xlsx``."""
    year, m = month.split("-")
    return f"{prime_id}_settlement_{_MONTH_NAMES[int(m) - 1]}_{year}.xlsx"


def build_xlsx(prime_id: str, month: str) -> Path:
    cell_dir = _REPO / "settlements" / prime_id / month
    prov     = _read_provenance(cell_dir)
    sheet, _totals = compute_sheet_rows(prov, prime_id)
    cfg      = _read_prime_yaml(prime_id)
    sde      = _read_sde(prime_id, date.fromisoformat(prov["period"]["start"]))

    wb = Workbook()
    _write_summary(wb.active, prov, sheet)
    _write_venues(wb.create_sheet(), sheet, cfg)
    _write_sky_revenue(wb.create_sheet(), prov, sheet, cfg)
    _write_sde(wb.create_sheet(), sde, sheet)
    if prov.get("sky_revenue_daily"):
        # "Debt" debug tab — emitted whenever the daily series is captured
        # (every run from the version that introduced ``sky_revenue_daily``
        # in provenance.json). Skipped silently on legacy outputs where
        # the field is absent.
        _write_debt(wb.create_sheet(), prov)
    if prov.get("display_only_breakdown"):
        # Only emit the off-protocol-holdings tab when the prime has display-
        # only venues this period. Avoids an empty tab cluttering monthly
        # reports for primes with no off-protocol positions.
        _write_off_protocol_holdings(wb.create_sheet(), prov, cfg)
    if any(b.get("burn_date") for b in (prov.get("sde_daily_breakdown") or [])):
        # Only emit the "SDE daily" tab when at least one SDE venue had a
        # burn this period — otherwise the three-column decomposition is
        # uninteresting (Sky = full position, Grove = 0 throughout).
        _write_sde_daily(wb.create_sheet(), prov)

    out = cell_dir / _output_filename(prime_id, month)
    wb.save(out)
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--prime", default="grove")
    parser.add_argument("--month", default="2026-04")
    args = parser.parse_args()
    out = build_xlsx(args.prime, args.month)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
