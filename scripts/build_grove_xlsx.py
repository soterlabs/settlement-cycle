"""Render a Grove-spreadsheet-shaped multi-tab xlsx from a settlement output.

Companion to ``build_monthly_report.py`` (which writes a `grove_sheet.md` +
`grove_sheet.csv`). This script reads those same artifacts plus the
prime-config files and emits a single ``grove_sheet.xlsx`` with several
tabs that mirror the layout of Grove's PnL workbook so the recipient can
diff side-by-side without a parser.

Tabs:
  * Headline       — month totals + reconciliation identities
  * Summary Comp   — per-venue Position / P2S / P2G / Aggregate columns
                     (matches Grove's "Summary Comp" sheet layout)
  * Grove Exposures — static venue config (label, chain, category, SDE)
  * Methodology    — bullet list of known methodology differences

Tabs that Grove's workbook has but **we cannot generate from current pipeline
output** without code changes (the pipeline doesn't emit daily-resolution
per-venue data today):
  * Asset-Level PnL — daily per-venue value + daily CoF allocation
  * USDS Line      — daily debt / utilized / SDE-deduction / agent_rate
  * JAAA_ETH Allocation — daily capped sd_share table
  * Transactions   — tx-level transfer feed
  * Holdings       — daily per-venue balance/NAV snapshot
  * Rewards        — tx-level Merkl claims
  * Treasury Rate  — subsidy reference rate daily

Stub placeholders are written for each so the file structure mirrors
Grove's; populate them in a follow-up PR by emitting the relevant
timeseries from ``compute_monthly_pnl``.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parent.parent

# Tag styles used across tabs.
_BOLD = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="E0E7EF")
_USD_FMT = '"$"#,##0.00;"−$"#,##0.00;"$"0.00'
_PCT_FMT = '0.00%'


def _D(x) -> Decimal:
    if x is None or x == "":
        return Decimal("0")
    if isinstance(x, Decimal):
        return x
    return Decimal(str(x))


def _set_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _bold_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL


def _read_grove_sheet_csv(cell_dir: Path) -> list[dict]:
    rows: list[dict] = []
    with (cell_dir / "grove_sheet.csv").open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows


def _read_provenance(cell_dir: Path) -> dict:
    with (cell_dir / "provenance.json").open(encoding="utf-8") as f:
        return json.load(f)


def _read_prime_config(prime_id: str) -> dict:
    p = _REPO / "config" / f"{prime_id}.yaml"
    with p.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def _read_sde(prime_id: str, period_start: date) -> dict[str, dict]:
    with (_REPO / "config" / "sky_direct_exposures.yaml").open(encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    out: dict[str, dict] = {}
    for section in ("active", "historical"):
        for e in cfg.get(section, []) or []:
            if e.get("prime") != prime_id or e.get("kind") == "pattern":
                continue
            start = e["start_date"]
            end = e.get("end_date")
            if isinstance(start, str): start = date.fromisoformat(start)
            if isinstance(end, str):   end   = date.fromisoformat(end)
            if start > period_start or (end is not None and end < period_start):
                continue
            vid = e["venue_id"]
            out[vid] = {"kind": e["kind"], "cap_usd": e.get("cap_usd"), "label": e.get("label", "")}
    return out


# --------------------------------------------------------------------------
# Tab writers
# --------------------------------------------------------------------------

def _write_headline(ws, prov: dict, sheet_rows: list[dict]) -> None:
    par = _D(prov["results"]["prime_agent_revenue"])
    ar  = _D(prov["results"]["agent_rate"])
    sky = _D(prov["results"]["sky_revenue"])
    sd  = sum((_D(r["sd_revenue"]) for r in sheet_rows), Decimal("0"))
    cof = sky - sd
    p2g_sum = sum((_D(r["profit_to_grove"]) for r in sheet_rows), Decimal("0"))

    sky_gross    = _D(prov["results"].get("sky_revenue_gross") or 0)
    spread_reimb = _D(prov["results"].get("susds_spread_reimbursement") or 0)

    ws.title = "Headline"
    ws.append([f"Grove monthly settlement — {prov['month']}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Component", "USD"])
    _bold_header(ws, 3, 2)
    rows = [
        ("Σ Profit to Sky ≡ sky_revenue",                                 sky),
        ("    ↳ CoF on Net_Subs (BR × utilized)",                         cof),
        ("    ↳ SDE revenue (full flow to Sky)",                          sd),
        ("    ↳ sUSDS spread reimb. (−Sky Revenue)",                      -spread_reimb),
        ("",                                                              None),
        ("Sky Revenue (max) — BR × full ilk debt, no deductions",          sky_gross),
        ("    ↳ CoF on Net_Subs (actual BR × utilized)",                  cof),
        ("    ↳ reduction from idle/SDE deductions (est., known venues)", -(sky_gross - cof) if sky_gross > 0 else Decimal("0")),
        ("",                                                              None),
        ("Σ Grove Net Payment (= prime_agent_revenue − CoF)",             p2g_sum),
        ("    ↳ prime_agent_revenue (per-venue gross venue yield total)", par),
        ("    ↳ CoF deducted by Grove (= CoF above)",                     -cof),
        ("agent_rate (subproxy yield, off-sheet)",                        ar),
        ("",                                                              None),
        ("Reconciliation drift Σ P2S − sky_revenue (must be ~0)",         sum((_D(r["profit_to_sky"]) for r in sheet_rows), Decimal("0")) - sky),
        ("Reconciliation drift (Σ GNP + CoF) − prime_agent_revenue",      p2g_sum + cof - par),
    ]
    for label, val in rows:
        ws.append([label, float(val) if val is not None else None])
        if val is not None:
            ws.cell(ws.max_row, 2).number_format = _USD_FMT

    ws.append([])
    ws.append(["Period", f"{prov['period']['start']} → {prov['period']['end']} ({prov['period']['n_days']} days)"])
    ws.append(["Generated at (UTC)", prov.get("generated_at_utc", "")])
    ws.append(["Pipeline version",   prov.get("settle_version", "")])

    ws.append([])
    note_row = ws.max_row + 1
    ws.append([
        "⚠ Note on Sky Revenue (max): this figure is BR × full ilk debt and is "
        "not a true ceiling on actual sky_revenue. SDE revenue (Σ sd_revenue) is "
        "added to sky_revenue on top of the BR charge, so actual sky_revenue can "
        "exceed this 'max' for primes with significant SDE positions. The subsidy "
        "(ref_rate ramp) is already applied — the rate used here is the same "
        "subsidised BR as in actual sky_revenue, not the raw Maker base rate. "
        "The figure is useful for seeing how much the idle-USDS / SDE / lending "
        "deductions reduce the BR component, and for per-venue Sky Rev Reduction "
        "estimates in Summary Comp (spread_reimb exact; utilized-deduction portion "
        "estimated proportionally from avg deduction; PSM3/Curve deductions not per-venue)."
    ])
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 72
    ws.cell(note_row, 1).font = Font(italic=True, color="666666")

    _set_widths(ws, {1: 80, 2: 22})


def _write_summary_comp(ws, sheet_rows: list[dict]) -> None:
    ws.title = "Summary Comp"
    cols = ["Venue ID", "Label", "Avg Value", "Weight", "Profit to Sky",
            "Revenue", "Grove Net Payment", "CoF Allocation", "SDE Revenue",
            "Spread Reimb", "Utilized Deduction (avg)", "Sky Rev Reduction (est.)",
            "Position (SoM)", "Position (EoM)", "Notes"]
    ws.append(cols)
    _bold_header(ws, 1, len(cols))

    # Sort by Profit to Sky descending
    for r in sorted(sheet_rows, key=lambda x: float(x["profit_to_sky"]), reverse=True):
        ws.append([
            r["venue_id"],
            r["label"],
            float(r["avg_value"]),
            float(r["weight"]),
            float(r["profit_to_sky"]),
            float(r["revenue"]),
            float(r["profit_to_grove"]),
            float(r["cof_alloc"]),
            float(r["sd_revenue"]),
            float(r["spread_reimb"]),
            float(r.get("deduction_avg") or 0),
            float(r.get("sky_rev_reduction_est") or 0),
            float(r["value_som"]),
            float(r["value_eom"]),
            r.get("note", ""),
        ])
    # Apply formats — cols: 3=AvgVal, 5=P2S, 6=Rev, 7=GNP, 8=CoF, 9=SDE,
    #   10=SpreadReimb, 11=Deduction, 12=SkyRevRedEst, 13=SoM, 14=EoM
    for row in range(2, ws.max_row + 1):
        for c in (3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14):
            ws.cell(row, c).number_format = _USD_FMT
        ws.cell(row, 4).number_format = _PCT_FMT

    _set_widths(ws, {1: 8, 2: 55, 3: 17, 4: 9, 5: 17, 6: 17, 7: 17, 8: 17,
                     9: 17, 10: 17, 11: 20, 12: 22, 13: 17, 14: 17, 15: 40})


def _write_grove_exposures(ws, prime_cfg: dict, sde_active: dict) -> None:
    ws.title = "Grove Exposures"
    ws.append(["Venue ID", "Label", "Chain", "Category", "Token", "SDE Status",
               "SDE Kind", "SDE Cap (USD)", "Pricing Notes"])
    _bold_header(ws, 1, 9)
    for v in prime_cfg.get("venues", []):
        vid = v.get("id", "")
        sde = sde_active.get(vid)
        cat = v.get("pricing_category", "")
        chain = v.get("chain", "")
        tok = v.get("token", {}) or {}
        ws.append([
            vid,
            v.get("label", ""),
            chain,
            cat,
            tok.get("symbol", ""),
            "ACTIVE" if sde else "—",
            (sde or {}).get("kind", ""),
            (sde or {}).get("cap_usd", "") or "",
            v.get("nav_oracle", {}).get("kind", "") if isinstance(v.get("nav_oracle"), dict) else "",
        ])
    _set_widths(ws, {1: 8, 2: 55, 3: 12, 4: 14, 5: 12, 6: 10, 7: 10, 8: 14, 9: 18})


def _write_methodology(ws, prov: dict) -> None:
    ws.title = "Methodology"
    ws.append(["MSC methodology notes (vs Grove workbook)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    blocks: list[tuple[str, str]] = [
        ("1. CoF charged at ilk level (not per venue)",
         "Sky's BR charge is applied once on Σ_d max(utilized_d, 0) × daily_factor, where "
         "utilized_d = cum_debt_d − idle_USDS_d − PSM_USDS_d − Σ SDE_asset_value_d − Curve_idle_d − lending_idle_d. "
         "The per-venue CoF allocation in 'Summary Comp' is a post-hoc display re-attribution by "
         "avg_value × (1 − sd_share); the canonical sky_revenue is the aggregate."),
        ("2. Merkl rewards booked on the on-chain claim date",
         "Cat C aTokens (E1 aHorRwaRLUSD, E3 aEthRLUSD) receive Merkl distributions as aToken receipts "
         "to the ALM. We extract them via a Dune Claimed×Mint event JOIN and book them in the month the "
         "claim transaction landed. Q1+Apr 2026 claims: Feb 6 ≈$3.78M, Apr 24 ≈$2.39M. Grove's sheet "
         "accrues unclaimed rewards across the period instead."),
        ("3. APY composition is multiplicative",
         "Base rate = combine_apys(SSR, 30bps) = (1+SSR)(1+0.003) − 1. Grove's 'Adj CoF' cell evidences "
         "additive composition (Sky CoF = SSR + 30bps = 0.043 in Jan, Adj CoF = ln(1.043)). The two "
         "differ by ~1.2 bps at SSR=4%, ≈$15K/month on Grove's ~$1.4B utilized."),
        ("4. SDE per Atlas — fixed and capped kinds",
         "Active for Grove: BUIDL (E10, fixed, since 2025-10-30), JTRSY (E9, fixed, since 2025-10-30). "
         "Historical: JAAA-ETH (E8, capped at $325M, 2025-10-23 → 2026-03-12). Daily sd_share_d = "
         "min(cap, v_d)/v_d for capped; sd_revenue_v = actual_revenue_v − revenue_v + external_revenue_v."),
        ("5. Subsidy ramp",
         "ref_rate (T-bill 3m) + (base_apy − ref_rate) × min(T, 24) / 24. For Grove 2026 the cap_usd "
         "is $1B and program_start is 2026-01-01."),
        ("6. Position values match Grove sheet within 0.1% for 14/16 mapped venues",
         "Exceptions: GACLO-1 (E21) — we model via per-venue cash_distributions, no daily NAV tracking; "
         "UNIV3 (E12) — the V3 LP wasn't on-chain until Feb 2026, we book $0 in Jan; Grove pre-books a "
         "$25M notional."),
    ]
    for header, body in blocks:
        ws.append([header])
        ws.cell(ws.max_row, 1).font = _BOLD
        ws.append([body])
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 60
        ws.append([])
    _set_widths(ws, {1: 120})


def _write_stub(ws, name: str, description: str) -> None:
    ws.title = name
    ws.append([f"[STUB] {name}"])
    ws["A1"].font = Font(bold=True, size=14, color="888888")
    ws.append([])
    ws.append([description])
    ws.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 90
    _set_widths(ws, {1: 120})


def build_xlsx(prime_id: str, month: str) -> Path:
    cell_dir = _REPO / "settlements" / prime_id / month
    prov     = _read_provenance(cell_dir)
    rows     = _read_grove_sheet_csv(cell_dir)
    cfg      = _read_prime_config(prime_id)
    sde      = _read_sde(prime_id, date.fromisoformat(prov["period"]["start"]))

    wb = Workbook()
    _write_headline(wb.active, prov, rows)
    _write_summary_comp(wb.create_sheet(), rows)
    _write_grove_exposures(wb.create_sheet(), cfg, sde)
    _write_methodology(wb.create_sheet(), prov)

    # Stub tabs for parity with Grove's layout — populated by a future PR
    # once the pipeline emits the corresponding daily-resolution feeds.
    _write_stub(
        wb.create_sheet(), "Asset-Level PnL",
        "Daily per-venue value + daily CoF allocation. The pipeline currently "
        "writes only SoM/EoM snapshots to venues.csv. Populating this tab "
        "requires emitting a daily timeseries per venue from compute_monthly_pnl "
        "(the data is computed internally but not serialized).",
    )
    _write_stub(
        wb.create_sheet(), "USDS Line",
        "Daily debt, utilized, SDE deductions, and agent rate. The pipeline "
        "computes these per day inside compute_sky_revenue but doesn't export "
        "the per-day frame today.",
    )
    _write_stub(
        wb.create_sheet(), "JAAA_ETH Allocation",
        "Daily capped sd_share table for E8. The compute layer iterates this "
        "daily (see _daily_capped_sd_revenue in prime_agent_revenue.py); "
        "exposing the per-day frame is a small additive change.",
    )
    _write_stub(
        wb.create_sheet(), "Transactions",
        "Tx-level transfer feed. Available from Dune raw_data (transfer_timeseries.sql) "
        "but not currently joined to the settlement artifacts.",
    )
    _write_stub(
        wb.create_sheet(), "Holdings",
        "Daily per-venue balance + NAV snapshot. Same constraint as Asset-Level "
        "PnL — daily data computed but not serialized.",
    )
    _write_stub(
        wb.create_sheet(), "Rewards",
        "Tx-level Merkl claims. The Dune query merkl_claims_ethereum.sql "
        "captures the (claim_tx, venue, amount) tuples. Wiring these into the "
        "settlement xlsx as a separate tab is a small follow-up.",
    )
    _write_stub(
        wb.create_sheet(), "Treasury Rate",
        "Daily subsidy reference rate (T-bill 3m). Source: "
        "config/subsidy_reference_rates.yaml. Could be exported as-is.",
    )

    out_path = cell_dir / "grove_sheet.xlsx"
    wb.save(out_path)
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--prime", default="grove")
    parser.add_argument("--month", default="2026-04")
    args = parser.parse_args()
    out = build_xlsx(args.prime, args.month)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
