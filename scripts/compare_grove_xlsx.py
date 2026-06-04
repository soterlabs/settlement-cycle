"""Compare our per-venue ``actual_revenue`` against Grove's reference xlsx.

Reads:
  data/grove/{jan,feb,mar,apr}_2026.xlsx       ← Grove's "Summary Comp" sheet
  settlements/grove/{month}/provenance.json    ← our pipeline output

Maps Grove tickers (e.g. ACRDX_PLUME) to our venue ids (E22) and prints a
side-by-side table per month with the residual gap.

Use to surface remaining methodology gaps after re-running the pipeline.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

# Grove ticker → our venue id(s). Use "+" to indicate we sum across multiple
# venue rows (the alt-holder pattern, etc.) before comparing.
TICKER_TO_VENUE: dict[str, str] = {
    "ACRDX_PLUME":           "E22",
    "aEthRLUSD_ETH":         "E3",
    "aHorRwaRLUSD_ETH":      "E1",
    "aHorRwaUSDC_ETH":       "E2",
    "BUIDLI_ETH":            "E10",
    "GACLO1_AVAX":           "E21",
    "grove-bbqAUSD_MON":     "E25",
    # Grove tracks two on-chain bbqAUSD vaults on Ethereum:
    #   grove-bbqAUSD_ETH     — older Morpho V1 vault (we intentionally don't track)
    #   grove-bbqAUSD-V2_ETH  — newer Morpho V2 vault — this is our E6
    # Verified by symbol() and per-month numbers (E6 actual_rev matches the V2
    # column to within rounding: Feb $5,696 ≈ $5,754 / Mar $40,044 ≈ $40,059 /
    # Apr $49,855 ≈ $49,837).
    "grove-bbqAUSD-V2_ETH":  "E6",
    "grove-bbqUSDC_BASE":    "E19",
    "grove-bbqUSDC_ETH":     "E5",
    "grove-bbqUSDC-V2_ETH":  "E4",
    "JAAA_AVAX":             "E20",
    "JAAA_ETH":              "E8",
    "JAAA_ETH_Grove":        "E8_GROVE",  # Mar-only — Grove splits JAAA into Sky/Grove parts
    "JAAA_ETH_Sky":          "E8_SKY",
    "JTRSY_ETH":             "E9",
    "STAC_ETH":              "E7",
    "UNIV3-AUSD-USDC_ETH":   "E12+E30",
    "CURVE-AUSD-USDC_ETH":   "E11",
    "USDC_ETH":              "E15+E32",
    "AUSD_ETH":              "E14+E31",
    "RLUSD_ETH":             "E13",
}

_REPO = Path(__file__).resolve().parent.parent


def _grove_per_venue(month_xlsx: Path) -> dict[str, tuple[float, float, float, float]]:
    """Return ``{ticker: (sky, grove, aggregate_with_rewards, aggregate_without_rewards)}``.

    Grove's Summary Comp sheet has TWO Aggregate columns side by side:

    * cols 3-5: Profit to Sky / Profit to Grove / Aggregate  (WITH rewards)
    * cols 6-8: Profit to Sky / Profit to Grove / Aggregate  (WITHOUT rewards)

    The "with rewards" aggregate folds in Merkl / Anchorage drops on an
    accrual basis (Grove allocates them ratably across the period); the
    "without rewards" aggregate is pool-native yield only. Our pipeline's
    ``actual_revenue`` realizes external rewards at CLAIM date in a
    separate ``external_revenue`` bucket — so the apples-to-apples Grove
    column for ``actual_revenue`` is the **without-rewards** Aggregate.
    The with-rewards column is shown alongside so the Merkl methodology
    delta is visible at a glance.
    """
    wb = openpyxl.load_workbook(month_xlsx, data_only=True)
    ws = wb["Summary Comp"]
    out: dict[str, tuple[float, float, float, float]] = {}
    for row_idx in range(4, ws.max_row + 1):
        ticker = ws.cell(row_idx, 2).value
        if ticker is None or (isinstance(ticker, str) and ticker.strip() == ""):
            break
        try:
            sky = float(ws.cell(row_idx, 3).value or 0)
            grove = float(ws.cell(row_idx, 4).value or 0)
            agg_with = float(ws.cell(row_idx, 5).value or 0)
            agg_without = float(ws.cell(row_idx, 8).value or 0)
        except (TypeError, ValueError):
            continue
        out[str(ticker)] = (sky, grove, agg_with, agg_without)
    return out


def _our_per_venue(month: str) -> dict[str, float]:
    out: dict[str, float] = {}
    path = _REPO / "settlements" / "grove" / month / "provenance.json"
    with path.open() as f:
        prov = json.load(f)
    for r in prov["venue_breakdown"]:
        out[r["venue_id"]] = float(r["actual_revenue"])
    return out


def compare(month: str, xlsx_name: str, *, gap_threshold: float = 5000) -> None:
    grove = _grove_per_venue(_REPO / "data" / "grove" / xlsx_name)
    ours = _our_per_venue(month)
    print(f"\n=== {month} — per-venue Aggregate (ours vs Grove) ===")
    print(f"{'Ticker':<25}{'Venue':<14}{'Ours actual_rev':>17}"
          f"{'Grove (no rew)':>15}{'Δ vs no-rew':>13}"
          f"{'Grove (w/ rew)':>17}  Note")
    print("-" * 110)

    # Stable order: largest |Grove no-rewards aggregate| first.
    seen_vids: set[str] = set()
    for ticker, (sky, grove_, agg_with, agg_without) in sorted(
        grove.items(), key=lambda kv: -abs(kv[1][3])
    ):
        vid = TICKER_TO_VENUE.get(ticker, "?")
        if vid == "?":
            print(f"{ticker:<25}{'?':<14}{'(no mapping)':>17}"
                  f"{agg_without:>15,.0f}{'':>13}{agg_with:>17,.0f}")
            continue
        # JAAA Mar split — combine the two halves before comparing.
        if vid == "E8_SKY":
            continue
        if vid == "E8_GROVE":
            paired = grove.get("JAAA_ETH_Sky", (0, 0, 0, 0))
            agg_with += paired[2]
            agg_without += paired[3]
            vid = "E8"
        for v in vid.split("+"):
            seen_vids.add(v)
        ids = vid.split("+")
        our_val = sum(ours.get(i, 0.0) for i in ids if i in ours)
        diff_without = our_val - agg_without
        flag = " ⚠" if abs(diff_without) > gap_threshold else ""
        note = ""
        if "+" in vid:
            note = f"  (sum across {vid})"
        # Merkl methodology delta = "with rewards" − "without rewards".
        # Mention it for tickers where the two differ materially so the
        # operator sees that the with-rewards gap is methodology-known.
        if abs(agg_with - agg_without) > 1000 and not note:
            note = f"  (Merkl accrual Δ ≈ ${agg_with - agg_without:,.0f})"
        print(f"{ticker:<25}{vid:<14}{our_val:>17,.0f}"
              f"{agg_without:>15,.0f}{diff_without:>13,.0f}{flag}"
              f"{agg_with:>17,.0f}{note}")

    extras = [vid for vid in ours if vid not in seen_vids and abs(ours[vid]) > 100]
    if extras:
        print("\n  Venues we report that Grove doesn't list:")
        for vid in extras:
            print(f"    {vid:<8}  actual_revenue = ${ours[vid]:>14,.0f}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--months", default="2026-01,2026-02,2026-03")
    args = parser.parse_args()
    months_xlsx = {
        "2026-01": "jan_2026.xlsx",
        "2026-02": "feb_2026.xlsx",
        "2026-03": "mar_2026.xlsx",
        "2026-04": "apr_2026.xlsx",
    }
    for m in args.months.split(","):
        m = m.strip()
        if m in months_xlsx:
            compare(m, months_xlsx[m])
        else:
            print(f"(no xlsx mapping for {m})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
