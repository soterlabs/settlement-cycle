#!/usr/bin/env python3
"""Extract Soter-side reconciliation inputs.

Sources (configured in ../sources.json; URLs primary, local fallback):
  - DR workbook (dr_comparison_latest.xlsx): grouped per prime via the authoritative
    `group` column on the `Summary` sheet. Soter DR = `Soter by Ref Code`; as-paid
    baseline = `Payouts`. Both rolled up by the same ref_code -> group map.
  - MSC primitives from settlements/<prime>/<month>/summary.md
    (agent_rate, prime_agent_profit, sky_revenue).

Writes soter_data.json next to forum_data.json.
Run:  python3 extract_soter.py
"""
import io, json, os, re, sys, urllib.request, urllib.error
from pathlib import Path

ROOT = Path(__file__).resolve().parent                 # reconciliation/
NEBULA = Path(__file__).resolve().parents[1]            # .../nebula
SRC = json.loads((ROOT / "sources.json").read_text())
FORUM_DATA = Path(os.environ.get("FORUM_DATA", ROOT / "forum_data.json"))
OUT = Path(os.environ.get("SOTER_DATA", ROOT / "soter_data.json"))

PRIME_DIR = SRC["prime_dir"]                            # display -> settlements dir
SHEETS = SRC["dr_xlsx"]["sheets"]
BASELINE_KEY = SRC["dr_xlsx"]["dr_baseline_sheet"]      # e.g. "payouts"

def _fetch_bytes(url, local):
    try:
        with urllib.request.urlopen(url, timeout=30) as r:
            print(f"  fetched {url}")
            return r.read()
    except Exception as e:
        if local and Path(local).exists():
            print(f"  URL failed ({e}); using local {local}")
            return Path(local).read_bytes()
        raise SystemExit(f"FATAL: could not fetch {url} ({e}) and no local fallback")

def _fetch_text(url):
    try:
        with urllib.request.urlopen(url, timeout=30) as r:
            return r.read().decode()
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None                                    # genuinely absent
        print(f"  WARNING: HTTP {e.code} for {url} — falling back to local")
        return None
    except Exception as e:                                 # timeout / DNS / reset / 5xx
        print(f"  WARNING: fetch failed for {url} ({e}) — falling back to local")
        return None

def _num(s):
    if s is None:
        return None
    s = str(s).replace("**", "").replace(",", "").strip()
    if s in ("", "TBD", "—", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def _refcode(c):
    if c is None:
        return None
    if isinstance(c, (int, float)):
        return int(c)
    m = re.match(r"\s*(-?\d+)", str(c))
    return int(m.group(1)) if m else None

# ---- DR workbook ----------------------------------------------------------
def load_dr():
    import openpyxl
    cfg = SRC["dr_xlsx"]
    data = _fetch_bytes(cfg["url"], NEBULA / cfg["local_fallback_relpath"])
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    def sheet(name):
        if name not in wb.sheetnames:
            raise SystemExit(f"FATAL: sheet {name!r} not in workbook {wb.sheetnames} "
                             f"— update sources.json dr_xlsx.sheets")
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            raise SystemExit(f"FATAL: sheet {name!r} is empty")
        return rows

    # 1) build ref_code -> group from the Summary sheet (group label carries down;
    #    skip 'Total' and blank rows).
    rows = sheet(SHEETS["summary"])
    refgroup, cur = {}, None
    for r in rows[1:]:
        g, rc = r[0], r[1]
        if g is not None:
            cur = str(g)
        if rc is None or str(rc).strip().lower() == "total":
            continue
        code = _refcode(rc)
        if code is not None and cur:
            refgroup[code] = cur

    def rollup(sheet_name):
        rows = sheet(sheet_name)
        hdr = rows[0]
        midx = {h: i for i, h in enumerate(hdr)
                if isinstance(h, str) and re.match(r"\d{4}-\d{2}$", h)}
        agg, unmapped = {}, set()
        for r in rows[1:]:
            code = _refcode(r[0])
            if code is None:
                continue
            grp = refgroup.get(code)
            if grp is None:
                unmapped.add(code)
                grp = "Other"
            for mon, i in midx.items():
                v = r[i]
                if isinstance(v, (int, float)):
                    agg.setdefault(grp, {}).setdefault(mon, 0.0)
                    agg[grp][mon] += float(v)
        for grp in agg:
            for mo in agg[grp]:
                agg[grp][mo] = round(agg[grp][mo], 2)
        return agg, sorted(unmapped)

    dr_soter, un1 = rollup(SHEETS["soter"])
    dr_base, un2 = rollup(SHEETS[BASELINE_KEY])
    if un1 or un2:
        print(f"  note: ref codes not in Summary group map -> 'Other': "
              f"{sorted(set(un1) | set(un2))}")
    return dr_soter, dr_base, SHEETS[BASELINE_KEY]

# ---- MSC summaries --------------------------------------------------------
HEADLINE = {"agent rate": "agent_rate",
            "prime agent profit": "prime_agent_profit",
            "sky revenue": "sky_revenue"}

def parse_summary(txt):
    out = {}
    for line in txt.splitlines():
        m = re.match(r"\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*$", line)
        if not m:
            continue
        label = m.group(1).replace("**", "").strip().lower()
        if label in HEADLINE:
            v = _num(m.group(2))
            if v is not None:
                out[HEADLINE[label]] = v
    return out

def load_msc(window):
    cfg = SRC["settlements"]
    base_url = os.environ.get("SETTLEMENTS_BASE_URL", cfg["base_url"])
    local_base = Path(os.environ.get(
        "SETTLEMENTS_DIR", NEBULA / cfg["local_fallback_relpath"]))
    profit, sky, ar = {}, {}, {}
    missing = []
    for prime, d in PRIME_DIR.items():
        profit[prime], sky[prime], ar[prime] = {}, {}, {}
        for m in window:
            rel = f"{d}/{m}/summary.md"
            txt = _fetch_text(f"{base_url}/{rel}")
            if txt is None:
                f = local_base / rel
                txt = f.read_text() if f.exists() else None
            if txt is None:
                missing.append(rel)
                continue
            vals = parse_summary(txt)
            if "prime_agent_profit" in vals: profit[prime][m] = vals["prime_agent_profit"]
            if "sky_revenue" in vals:        sky[prime][m] = vals["sky_revenue"]
            if "agent_rate" in vals:         ar[prime][m] = vals["agent_rate"]
    if missing:
        print(f"  WARNING: no summary.md (URL or local) for {len(missing)} prime-month(s): "
              f"{', '.join(missing)} — these rows will be BLANK in the report, not $0")
    return profit, sky, ar

def main():
    fd = json.loads(FORUM_DATA.read_text())
    window = fd["window_months"]
    print("DR workbook:")
    dr_soter, dr_base, base_name = load_dr()
    print("MSC summaries:")
    profit, sky, ar = load_msc(window)

    # The report looks DR up by PRIME_DIR display name; DR is keyed by the workbook's
    # `group` label. Surface any prime whose display name has no matching DR group
    # (expected for genuinely DR-free primes like Obex; a DR-earner here = renamed group).
    dr_groups = set(dr_soter) | set(dr_base)
    no_dr = [p for p in PRIME_DIR if p not in dr_groups]
    if no_dr:
        print(f"  note: no DR group matches display name(s) {no_dr} — fine for non-DR "
              f"primes; if a DR-earning prime is listed, the workbook 'group' label was renamed")

    # DR in groups that map to no prime (e.g. 'Other', 'Osero') is dropped from the report.
    # That is correct only if those are genuinely non-prime (excluded venues). Warn loudly with
    # the magnitude so a mis-grouped prime's DR can never be silently omitted.
    nonprime = {g: v for g, v in dr_soter.items() if g not in PRIME_DIR}
    if nonprime:
        tot = sum(sum(v.values()) for v in nonprime.values())
        alltot = sum(sum(v.values()) for v in dr_soter.values()) or 1.0
        print(f"  WARNING: {tot:,.0f} of Soter DR ({tot/alltot:.0%} of total) is in non-prime "
              f"group(s) {sorted(nonprime)} — NOT attributed to any prime, excluded from the "
              f"reconciliation. Confirm these are excluded venues, not mis-grouped prime ref codes.")

    data = {
        "window_months": window,
        "primes": list(PRIME_DIR.keys()),
        "prime_agent_profit": profit,
        "sky_revenue": sky,
        "agent_rate": ar,
        "dr_soter": dr_soter,
        "dr_baseline": dr_base,
        "dr_baseline_name": base_name,
    }
    OUT.write_text(json.dumps(data, indent=2))
    print(f"\nwrote {OUT}")
    print(f"DR groups (Soter): {sorted(dr_soter)}")
    for p in PRIME_DIR:
        ms = sorted(profit.get(p, {}))
        print(f"  {p:8s} MSC months {ms}  DR-Soter {sorted(dr_soter.get(p, {}))}")

if __name__ == "__main__":
    sys.exit(main())
